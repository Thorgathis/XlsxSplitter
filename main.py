from random import randint
from openpyxl import Workbook
import os
import openpyxl


def checkFolderExists(name):
    if not os.path.exists(name):
        os.mkdir(name)
    else:
        name = name[:-1] + "-" + str(randint(1000, 9999)) + "/"
        os.mkdir(name)
    return name


# create folder for new files
foldername = checkFolderExists("new-data/")

# load work book
frame = openpyxl.load_workbook("book.xlsx")
act_frame = frame.active


headers = []
#get header from xlsx file
for col in act_frame.iter_cols(0):
    headers.append(str(col[0].value))

for row in range(1, act_frame.max_row):
    wb = Workbook()
    page = wb.active
    page.title = 'sheet1'  # set title to page

    data = []  # data to xlsx files
    ses_data = []  # row data

    for col in act_frame.iter_cols(1, act_frame.max_column):
        data.append(str(col[row].value))
        ses_data.append(col)

    page.append(headers)  # add headers to xlsx file
    page.append(data)  # add data to xlsx file

    wb.save(foldername + str(ses_data[0][row].value) + '.xlsx')  # save file with name from first col
